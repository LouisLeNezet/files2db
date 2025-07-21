#!/usr/bin/env python3
"""
Created on 22/10/2021
@author: LouisLeNezet
Check for presence absence of file, sheets, columns
"""

import logging
import re
from importlib import resources

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from ..read_file.data_read import read_file
from ..ui.get_infos import get_file_path


def load_file_orga():
    """Load internal file_orga.csv as a Pandas DataFrame and transform it into a dictionary."""
    with resources.open_text("files2db.data", "file_orga.csv") as csvfile:
        orga = pd.read_csv(csvfile)

        # Validate required columns
        required_columns = {"file", "columns_needed", "columns_sup"}
        missing_columns = required_columns - set(orga.columns)
        if missing_columns:
            raise KeyError(
                f"Missing required columns in organisation file: {sorted(missing_columns)}"
            )

        # Convert columns_needed from comma-separated strings to lists
        orga["columns_needed"] = orga["columns_needed"].apply(lambda x: x.split(","))

        # Convert columns_sup to boolean (if it's not already)
        orga["columns_sup"] = orga["columns_sup"].astype(bool)

        # Convert DataFrame to dictionary
        orga_dict = orga.set_index("file").to_dict(orient="index")

        return orga_dict


def validate_files_presence(files_needed: set, files_available: set, path: str):
    """Check if required files exist in the dataset."""
    missing_files = files_needed - files_available
    extra_files = files_available - files_needed

    if missing_files:
        raise KeyError(
            f"Missing files {', '.join(sorted(missing_files))} in my_organisation file {path}"
        )

    if extra_files:
        logging.warning(
            "Extra files %s present in %s and not needed", extra_files, path
        )


def validate_columns(df: pd.DataFrame, path: str, cols_need: list, cols_sup=False):
    """Check for missing and extra columns in each file."""
    missing = cols_need - set(df.columns)
    extra = set(df.columns) - cols_need

    if missing:
        raise KeyError(f"Missing columns {missing} in {path}")

    if not cols_sup and extra:
        logging.warning("Extra columns %s in %s and won't be used", extra, path)


def validate_columns_orga(orga_dict: dict, db_dict: dict):
    """Check for missing and extra columns in each file based on organisation specifications."""
    for file, df in db_dict.items():
        validate_columns(
            df,
            path=file,
            cols_need=set(orga_dict[file]["columns_needed"]),
            cols_sup=orga_dict[file]["columns_sup"],
        )
        # Transform to integer needed columns
        if ("integer" in orga_dict[file]) and isinstance(
            orga_dict[file]["integer"], str
        ):
            for col in orga_dict[file]["integer"].split(","):
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors="coerce").astype("Int64")
        # Transform to list needed columns
        if ("list" in orga_dict[file]) and isinstance(orga_dict[file]["list"], str):
            for col in orga_dict[file]["list"].split(","):
                if col in df.columns:
                    df[col] = df[col].apply(
                        lambda x: x.split(",") if isinstance(x, str) else x
                    )
    return db_dict


def get_db_from_excel(path: str, orga_dict: dict) -> dict:
    """Load and validate an Excel file based on organisation specifications."""
    path_file = get_file_path(path)

    try:
        wb = load_workbook(path_file, read_only=True)
    except FileNotFoundError as error:
        logging.error("File %s not found", path)
        raise error

    validate_files_presence(set(orga_dict.keys()), set(wb.sheetnames), path)

    db_dict = {sheet: pd.read_excel(path_file, sheet_name=sheet) for sheet in orga_dict}

    return db_dict


def get_db_from_csv(path: str, orga_dict: dict) -> dict:
    """Load and validate a CSV file based on organisation specifications."""
    path_file = get_file_path(path)
    all_db_files = pd.read_csv(path_file)

    if set(all_db_files.columns) != {"file", "path", "sep"}:
        raise KeyError(f"Columns in {path} should only contain ['file', 'path', 'sep']")

    validate_files_presence(set(orga_dict.keys()), set(all_db_files["file"]), path)

    db_dict = {}
    all_db_files.set_index("file", inplace=True)
    for file, file_path, sep in all_db_files.itertuples():
        db_path_file = get_file_path(file_path)
        db_dict[file] = read_file(db_path_file, sep=sep)

    return db_dict


def get_db_from_path(path_file: str, db_orga: dict) -> dict:
    """
    Get the database from a file based on its extension and organisation specifications.
    Parameters
    ----------
    path_file : str
        Full path to the file.
    db_orga : dict
        Organisation dictionary containing file specifications.
    Returns
    -------
    db_dict : dict
        Dictionary containing the database loaded from the file.
    Raises
    ------
    TypeError
        If the file extension is not supported (not .csv or .xlsx).
    """
    path_file = get_file_path(path_file)

    if not isinstance(path_file, str):
        raise TypeError("The path_file should be a string")
    if not isinstance(db_orga, dict):
        raise TypeError("The db_orga should be a dictionary")

    if re.search(string=path_file, pattern=r"\.csv$"):
        db_dict = get_db_from_csv(path_file, db_orga)
    elif re.search(string=path_file, pattern=r"\.(xlsx|xls|xlsm)"):
        db_dict = get_db_from_excel(path_file, db_orga)
    else:
        raise TypeError(
            f"File {path_file} should be either an .xlsx, .xls, xlsm or a .csv"
        )

    db_dict = validate_columns_orga(db_orga, db_dict)

    # Change all nan values to None
    for _, df in db_dict.items():
        df.replace(np.nan, None, inplace=True)

    return db_dict
